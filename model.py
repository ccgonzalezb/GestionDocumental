import pdfplumber
import pytesseract
from pdf2image import convert_from_path
import re
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

class PDFProcessor:
    def __init__(self, tesseract_cmd, excel_path):
        self.tesseract_cmd = tesseract_cmd
        self.excel_path = excel_path
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    def procesar_pdf(self, ruta_pdf):
        No = "" 
        FECHA = datetime.now().strftime("%Y-%m-%d")
        ORFEO = os.path.splitext(os.path.basename(ruta_pdf))[0]
        RAD_EXTERNO = ""
        FECHA_SALIDA = ""
        FECHA_DOC = ""
        ASUNTO = ""
        ORIGEN = ""
        CLASE_DOCUMENTO = ""
        SECEJ = ""
        JEMPP = ""
        COEJC = ""
        SECEJ2 = ""
        ORDEN_EMITIDA_JEMPP = ""
        DESTINO = ""
        CLASIFICACION_ORDEN = ""
        PLAZO = ""
        EMISOR_ORDEN = ""
        OBSERVACIONES = ""

        fecha_pattern = r'(\b\d{1,2}\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de\s+\d{4}\b)'
        radicado_pattern = [
                r'Radicado\s*N°[\s:]*([^\s\r:]*):?',
                r'Radicado\s*No\.[\s:]*([^\s\r:]*):?',
                r'Radicado\s*N.°[\s:]*([^\s\r:]*):?',
                r'Radicado\s*n\*[\s:]*([^\s\r:]*):?'
                r'Radicado\s*n°[\s:]*([^\s\r:]*):?'  
                r'Radicado[\s:]*([^\s\r:]*):?'                                                
                r'N°\s*Radicado[\s:]*([^\s\r:]*):?',
                r'NO\.\s*([^\s\r:]*):?',
                r'Nro\.\s*([^\s\r:]*):?',

            ]
        radicado_numeros_pattern = r'\b(\d{4}-\d{3}-\d{7}-\d{1}|\d{13,16}\/?)\b'
        palabras_detencion_observaciones = [r'\b(Brigadier|MG|BG|Coronel|Mayor|General)\b']
        frases_clave = [
            "DESPACHO COMANDANTE",
            "DESPACHO SEGUNDO COMANDANTE",
            "JEFATURA DE ESTADO MAYOR DE PLANEACIÓN Y POLÍTICAS",
            "JEFATURA DE ESTADO MAYOR GENERADOR DE FUERZA",
            "FUERZAS MILITARES DE COLOMBIA"
        ]

        with pdfplumber.open(ruta_pdf) as pdf:
            for numero_pagina, pagina in enumerate(pdf.pages):
                texto_pagina = pagina.extract_text()

                if not texto_pagina:
                    print(f"No se encontró texto en la página {numero_pagina + 1}. Aplicando OCR...")
                    imagen = convert_from_path(ruta_pdf, dpi=300, first_page=numero_pagina + 1, last_page=numero_pagina + 1)[0]
                    texto_pagina = pytesseract.image_to_string(imagen, lang="spa")

                lineas = texto_pagina.split('\n')

                if numero_pagina == 0:

                    for linea in lineas[:30]:
                        if FECHA_DOC == "":
                            fecha_match = re.search(fecha_pattern, linea, re.IGNORECASE)
                            if fecha_match:
                                FECHA_DOC = fecha_match.group(0)
                                print(f"FECHA_DOC encontrada: {FECHA_DOC}")
                                break

                    # Buscar RAD_EXTERNO
                    if RAD_EXTERNO == "":
                        for linea in lineas[:30]:
                            for pattern in radicado_pattern:
                                match = re.search(pattern, linea, re.IGNORECASE)
                                if match:
                                    radicado_encontrado = re.search(radicado_numeros_pattern, linea)
                                    if radicado_encontrado:
                                        valor_encontrado = radicado_encontrado.group(1)
                                        # Verificar si el valor encontrado es igual a ORFEO
                                        if valor_encontrado != ORFEO:
                                            RAD_EXTERNO = valor_encontrado.strip('/')  # Quita el '/' al final si existe
                                            print(f"Radicado válido encontrado: {RAD_EXTERNO}")
                                            break  # Sale del bucle de patrones
                                    # Si el valor coincide con ORFEO, sigue buscando otros patrones    
                            if RAD_EXTERNO:  # Si ya se asignó, rompe el bucle principal
                                #
                                break
                    # Buscar el Asunto solo en la primera página
                    asuntos = []
                    for linea in lineas[:30]:
                        if re.search(r'\b(Asunto|ASUNTO|asunto|Asunto:|ASUNT)\b', linea):
                            asunto_matches = re.findall(r'^\s*(Asunto|ASUNTO|asunto|Asunto:|ASUNT)\s*[:\-]?\s*(.*)', linea.strip())
                            for match in asunto_matches:
                                if match[1].strip():
                                    asuntos.append(match[1].strip())  # Añadir el asunto sin espacios extras
                    
                    # Si hay más de un "Asunto", tomar el segundo
                    if len(asuntos) > 1:
                        ASUNTO = asuntos[1]  # Segundo "Asunto" encontrado
                        print(f"Segundo Asunto encontrado: {ASUNTO}")
                    elif len(asuntos) == 1:
                        ASUNTO = asuntos[0]
                        print(f"Primer Asunto encontrado: {ASUNTO}")
                    
                    # Buscar CLASE_DOCUMENTO
                    if CLASE_DOCUMENTO == "":
                        for linea in lineas[:30]:

                            if re.search(r'\bRADIOGRAMA\b', texto_pagina):
                                CLASE_DOCUMENTO = "RADIOGRAMA"
                            elif re.search(r'\bPLAN\s+N\b', texto_pagina):
                                CLASE_DOCUMENTO = "PLAN"
                            elif re.search(r'\bCIRCULAR\s+N\b', texto_pagina):
                                CLASE_DOCUMENTO = "CIRCULAR"
                            elif re.search(r'\bBOLETÍN\b', texto_pagina):
                                CLASE_DOCUMENTO = "BOLETÍN"
                            else:
                                CLASE_DOCUMENTO = "HR"

                else:
                    capturar_observaciones = False
                    for linea in lineas[:10]:
                        if any(frase in linea for frase in frases_clave):
                            capturar_observaciones = True
                            print(f"Frase clave encontrada en página {numero_pagina + 1}: {linea}")
                            break

                    if capturar_observaciones:
                        for linea in lineas:
                            if re.search(r'\b(OBSERVACIONES)\b', linea, re.IGNORECASE):
                                indice_observaciones = lineas.index(linea)
                                observaciones = []
                                for i in range(indice_observaciones + 1, len(lineas)):
                                    if any(re.search(palabra, lineas[i]) for palabra in palabras_detencion_observaciones):
                                        break
                                    observaciones.append(lineas[i].strip())
                                if re.search(r'\bDESPACHO\s+COMANDANTE\b', texto_pagina, re.IGNORECASE):
                                    observaciones.insert(0, "COEJC:")
                                elif re.search(r'\bDESPACHO\s+SEGUNDO\s+COMANDANTE\b', texto_pagina, re.IGNORECASE):
                                    observaciones.insert(0, "SECEJ:")
                                elif re.search(r'\bJEFATURA\s+DE\s+ESTADO\s+MAYOR\s+DE\s+PLANEACIÓN\s+Y\s+POLÍTICAS\b', texto_pagina, re.IGNORECASE):
                                    observaciones.insert(0, "JEMPP:")   
                                ORDEN_EMITIDA_JEMPP += "\n".join(observaciones) + "\n"
                                print(f"Observaciones encontradas: {' '.join(observaciones)}")
                                if "SECEJ" in ORDEN_EMITIDA_JEMPP:
                                    EMISOR_ORDEN = "SECEJ"
                                elif "COEJC" in ORDEN_EMITIDA_JEMPP:
                                    EMISOR_ORDEN = "COEJC" 
                                else:
                                    EMISOR_ORDEN = "JEMPP"      
                                break
        data = {
            'No.': [No],
            'FECHA': [FECHA],
            'ORFEO': [ORFEO],
            'RAD_EXTERNO': [RAD_EXTERNO],
            'FECHA_SALIDA': [FECHA_SALIDA],
            'FECHA_DOC': [FECHA_DOC],
            'ASUNTO': [ASUNTO],
            'ORIGEN': [ORIGEN],
            'CLASE_DOCUMENTO': [CLASE_DOCUMENTO],
            'SECEJ': [SECEJ],
            'JEMPP': [JEMPP],
            'COEJC': [COEJC],
            'SECEJ2': [SECEJ2],
            'ORDEN_EMITIDA_JEMPP': [ORDEN_EMITIDA_JEMPP],
            'DESTINO': [DESTINO],
            'CLASIFICACION_ORDEN': [CLASIFICACION_ORDEN],
            'PLAZO': [PLAZO],
            'EMISOR_ORDEN': [EMISOR_ORDEN],
            'OBSERVACIONES': [OBSERVACIONES]
        }

        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(list(data.keys()))

        sheet.append([data[key][0] for key in data])
        wb.save(self.excel_path)
        print(f"Archivo Excel actualizado: {self.excel_path}")
